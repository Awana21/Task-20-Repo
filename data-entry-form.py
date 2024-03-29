#Task 20 (Updated Version)
#Issues Fixed: 
#Implemented Code Documentation
#Implemented proper error handling and providing feedback to the user

"""
Excel Viewer & Data Entry Form

This script implements both an Excel Viewer & Data Entry properties using tkinter library and openpyxl .
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl

# Function to load data from Excel file
def load_data():
    path = r"C:\Users\44784\Desktop\Slider\CoGrammar\Task 20\Python Data Entry\people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
# Existing code for loading data into the treeview
    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

# Function to insert a new row of data
def insert_row():
    name = name_entry.get()
    age = int(age_spinbox.get())
    subscription_status = status_combobox.get()
    employment_status = "Employed" if a.get() else "Unemployed"

    print(name, age, subscription_status, employment_status)

# Your existing code for inserting row into Excel sheet
    path = r"C:\Users\44784\Desktop\Slider\CoGrammar\Task 20\Python Data Entry\people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, employment_status]
    sheet.append(row_values)
    workbook.save(path)

# Your existing code for inserting row into treeview
    treeview.insert('', tk.END, values=row_values)

# Your existing code for clearing values
    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Age")
    status_combobox.set(combo_list[0])
    checkbutton.state(["!selected"])
=======
    try:
        path = r"C:\Users\44784\Desktop\Slider\CoGrammar\Task 20\Python Data Entry\people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        # Existing code for loading data into the treeview
        list_values = list(sheet.values)
        for col_name in list_values[0]:
            treeview.heading(col_name, text=col_name)
        for value_tuple in list_values[1:]:
            treeview.insert('', tk.END, values=value_tuple)
        workbook.close()  # Close the workbook after reading data
    except Exception as e:
        messagebox.showerror("Error", f"Error loading data: {str(e)}")

# Function to insert a new row of data
def insert_row():
    try:
        name = name_entry.get()
        age = int(age_spinbox.get())
        subscription_status = status_combobox.get()
        employment_status = "Employed" if a.get() else "Unemployed"

        # Your existing code for inserting row into Excel sheet
        path = r"C:\Users\44784\Desktop\Slider\CoGrammar\Task 20\Python Data Entry\people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [name, age, subscription_status, employment_status]
        sheet.append(row_values)
        workbook.save(path)
        workbook.close()  # Close the workbook after writing data

        # Your existing code for inserting row into treeview
        treeview.insert('', tk.END, values=row_values)

        # Your existing code for clearing values
        name_entry.delete(0, "end")
        name_entry.insert(0, "Name")
        age_spinbox.delete(0, "end")
        age_spinbox.insert(0, "Age")
        status_combobox.set(combo_list[0])
        checkbutton.state(["!selected"])

    except Exception as e:
        messagebox.showerror("Error", f"Error inserting row: {str(e)}")


# Function to toggle between light and dark themes
def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

# Main application window
root=tk.Tk()



# Widget Creation and GUI setup
style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")
style.theme_use("forest-dark")


frame = ttk.Frame(root)
frame.pack()

# Frame for inserting rows
widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

# Entry widget for name
name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Name")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

# Spinbox for age input
age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=100)
age_spinbox.insert(0, "Age")
age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew") 

# Combobox for subscription status
combo_list = ["Subscribed", "Not Subscribed", "Other"]

status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")

# Checkbox for employment status
a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=a)
checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

# Button to insert a new row
button = ttk.Button(widgets_frame, text="Insert", command=insert_row)
button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

# Separator and check button for theme mode
separator = ttk.Separator(widgets_frame)
separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(
    widgets_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

# Frame for displaying data in a treeview
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

# Columns for the treeview
cols = ("Name", "Age", "Subscription", "Employment")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Name", width=100)
treeview.column("Age", width=50)
treeview.column("Subscription", width=100)
treeview.column("Employment", width=100)
treeview.pack()
treeScroll.config(command=treeview.yview)

# Load data into the treeview
load_data()

# Run the main application loop
root.mainloop()


